"""
excel_builder.py
================
Generates a professionally formatted Excel workbook containing:

  Sheet 1: Dashboard         — Company overview, key metrics, scenario summary + charts
  Sheet 2: Income Statement  — Historical income statement data + margin chart
  Sheet 3: Balance Sheet     — Historical balance sheet data + structure chart
  Sheet 4: Cash Flow         — Historical cash flow data + waterfall chart
  Sheet 5: WACC              — WACC breakdown + capital structure pie + rate bar chart
  Sheet 6: DCF Base Case     — Detailed base-case DCF model + projection chart
  Sheet 7: DCF Bull Case     — Rising sales & profit, falling rates + projection chart
  Sheet 8: DCF Bear Case     — Falling sales & profit, rising rates + projection chart
  Sheet 9: DCF Rising Rates  — Stable growth, rate hikes + projection chart
  Sheet 10: DCF Falling Rates — Stable growth, rate cuts + projection chart
  Sheet 11: Scenario Summary — Side-by-side comparison + multi-chart page
  Sheet 12: Sensitivity      — WACC vs Terminal Growth sensitivity table
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, AreaChart, Reference
from openpyxl.chart.series import SeriesLabel, DataPoint
from openpyxl.chart.label import DataLabelList
from datetime import datetime


# ============================================================================
# STYLE CONSTANTS
# ============================================================================

# Colors
DARK_BLUE = "1B2A4A"
MED_BLUE = "2E5090"
LIGHT_BLUE = "D6E4F0"
ACCENT_GREEN = "27AE60"
ACCENT_RED = "E74C3C"
ACCENT_ORANGE = "F39C12"
ACCENT_TEAL = "0097A7"
WHITE = "FFFFFF"
LIGHT_GRAY = "F2F2F2"
MED_GRAY = "D9D9D9"
DARK_TEXT = "1A1A1A"
MONEY_GREEN = "E8F5E9"
MONEY_RED = "FFEBEE"

# Chart color palette (hex without #)
CHART_COLORS = ["2E5090", "27AE60", "E74C3C", "F39C12", "0097A7", "8E24AA", "1B2A4A"]

# Fonts
TITLE_FONT = Font(name="Calibri", size=18, bold=True, color=WHITE)
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color=WHITE)
SUBHEADER_FONT = Font(name="Calibri", size=11, bold=True, color=DARK_BLUE)
LABEL_FONT = Font(name="Calibri", size=10, color=DARK_TEXT)
VALUE_FONT = Font(name="Calibri", size=10, color=DARK_TEXT)
SMALL_FONT = Font(name="Calibri", size=9, color="666666")
LINK_FONT = Font(name="Calibri", size=10, color=MED_BLUE, underline="single")
BOLD_VALUE = Font(name="Calibri", size=10, bold=True, color=DARK_TEXT)
BIG_NUMBER = Font(name="Calibri", size=14, bold=True, color=DARK_BLUE)
GREEN_FONT = Font(name="Calibri", size=10, bold=True, color=ACCENT_GREEN)
RED_FONT = Font(name="Calibri", size=10, bold=True, color=ACCENT_RED)

# Fills
TITLE_FILL = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
HEADER_FILL = PatternFill(start_color=MED_BLUE, end_color=MED_BLUE, fill_type="solid")
LIGHT_FILL = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
ALT_ROW_FILL = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
GREEN_FILL = PatternFill(start_color=MONEY_GREEN, end_color=MONEY_GREEN, fill_type="solid")
RED_FILL = PatternFill(start_color=MONEY_RED, end_color=MONEY_RED, fill_type="solid")
WHITE_FILL = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")

# Borders
THIN_BORDER = Border(
    left=Side(style="thin", color=MED_GRAY),
    right=Side(style="thin", color=MED_GRAY),
    top=Side(style="thin", color=MED_GRAY),
    bottom=Side(style="thin", color=MED_GRAY),
)
BOTTOM_BORDER = Border(bottom=Side(style="medium", color=DARK_BLUE))

# Alignments
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Number formats
FMT_DOLLAR = '#,##0'
FMT_DOLLAR_M = '#,##0.0,,"M"'
FMT_DOLLAR_B = '#,##0.0,,,"B"'
FMT_PCT = '0.0%'
FMT_PCT2 = '0.00%'
FMT_PRICE = '$#,##0.00'
FMT_NUM = '#,##0'
FMT_RATIO = '0.00x'


# ============================================================================
# HELPERS
# ============================================================================

def _fmt_large(val):
    """Pick appropriate number format for large values."""
    if abs(val) >= 1e9:
        return FMT_DOLLAR_B
    elif abs(val) >= 1e6:
        return FMT_DOLLAR_M
    return FMT_DOLLAR


def _set_col_widths(ws, widths: dict):
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _write_title_row(ws, row, text, max_col=10):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = TITLE_FONT
    cell.fill = TITLE_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 36


def _write_header_row(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=start_col + i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def _write_data_row(ws, row, label, values, start_col=1, fmt=FMT_DOLLAR_B,
                    label_font=LABEL_FONT, value_font=VALUE_FONT, alt=False):
    cell = ws.cell(row=row, column=start_col, value=label)
    cell.font = label_font
    cell.alignment = LEFT
    cell.border = THIN_BORDER
    if alt:
        cell.fill = ALT_ROW_FILL

    for i, v in enumerate(values):
        cell = ws.cell(row=row, column=start_col + 1 + i, value=v)
        cell.font = value_font
        cell.alignment = RIGHT
        cell.border = THIN_BORDER
        if fmt:
            cell.number_format = fmt if not callable(fmt) else fmt(v)
        if alt:
            cell.fill = ALT_ROW_FILL


def _write_section_header(ws, row, text, max_col=10):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = SUBHEADER_FONT
    cell.fill = LIGHT_FILL
    cell.alignment = LEFT
    cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 22


def _write_kv(ws, row, col_label, label, col_val, value, fmt=None,
              label_font=LABEL_FONT, value_font=VALUE_FONT):
    """Write a label-value pair."""
    c1 = ws.cell(row=row, column=col_label, value=label)
    c1.font = label_font
    c1.alignment = LEFT
    c1.border = THIN_BORDER
    c2 = ws.cell(row=row, column=col_val, value=value)
    c2.font = value_font
    c2.alignment = RIGHT
    c2.border = THIN_BORDER
    if fmt:
        c2.number_format = fmt


def _style_chart(chart, width=28, height=14):
    """Apply consistent styling to a chart."""
    chart.width = width
    chart.height = height
    chart.legend.position = "b"


def _color_series(chart, colors=None):
    """Apply color palette to chart series."""
    colors = colors or CHART_COLORS
    from openpyxl.chart.shapes import GraphicalProperties
    from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
    for i, series in enumerate(chart.series):
        color = colors[i % len(colors)]
        series.graphicalProperties.solidFill = color


# ============================================================================
# CHART BUILDERS (reusable)
# ============================================================================

def _make_bar_chart(ws, title, y_title, data_rows, cat_row, data_start_col,
                    data_end_col, series_labels, anchor, width=28, height=14,
                    chart_type="col"):
    """Create and place a bar chart."""
    chart = BarChart()
    chart.type = chart_type
    chart.style = 10
    chart.title = title
    chart.y_axis.title = y_title
    _style_chart(chart, width, height)

    cats = Reference(ws, min_col=data_start_col, max_col=data_end_col, min_row=cat_row)
    for i, (row_idx, label) in enumerate(zip(data_rows, series_labels)):
        data = Reference(ws, min_col=data_start_col, max_col=data_end_col, min_row=row_idx)
        chart.add_data(data, from_rows=True, titles_from_data=False)
        chart.series[i].tx = SeriesLabel(v=label)
    chart.set_categories(cats)
    _color_series(chart)
    ws.add_chart(chart, anchor)
    return chart


def _make_line_chart(ws, title, y_title, data_rows, cat_row, data_start_col,
                     data_end_col, series_labels, anchor, width=28, height=14):
    """Create and place a line chart."""
    chart = LineChart()
    chart.style = 10
    chart.title = title
    chart.y_axis.title = y_title
    _style_chart(chart, width, height)

    cats = Reference(ws, min_col=data_start_col, max_col=data_end_col, min_row=cat_row)
    for i, (row_idx, label) in enumerate(zip(data_rows, series_labels)):
        data = Reference(ws, min_col=data_start_col, max_col=data_end_col, min_row=row_idx)
        chart.add_data(data, from_rows=True, titles_from_data=False)
        chart.series[i].tx = SeriesLabel(v=label)
        chart.series[i].graphicalProperties.line.width = 22000  # ~2pt
    chart.set_categories(cats)
    _color_series(chart)
    ws.add_chart(chart, anchor)
    return chart


# ============================================================================
# SHEET BUILDERS
# ============================================================================

def build_dashboard(wb, stock, financials, rates, model_result, data_source):
    """Sheet 1: Company overview dashboard with multiple charts."""
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_properties.tabColor = DARK_BLUE

    max_col = 10
    _set_col_widths(ws, {1: 22, 2: 18, 3: 5, 4: 22, 5: 18, 6: 5, 7: 22, 8: 18, 9: 5, 10: 18})

    # Title
    row = 1
    _write_title_row(ws, row, f"  DCF VALUATION MODEL  —  {stock['ticker']}", max_col)

    row = 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row=row, column=1,
                   value=f"  {stock['company_name']}  |  Data Source: {data_source.upper()}  |  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    cell.font = Font(name="Calibri", size=10, color=LIGHT_BLUE)
    cell.fill = TITLE_FILL
    ws.row_dimensions[row].height = 22

    # --- Company Info ---
    row = 4
    _write_section_header(ws, row, "COMPANY INFORMATION", max_col)

    info_items = [
        ("Ticker", stock["ticker"], None),
        ("Company", stock["company_name"], None),
        ("Sector", stock["sector"], None),
        ("Industry", stock["industry"], None),
        ("Country", stock["country"], None),
        ("Currency", stock["currency"], None),
    ]
    for i, (label, val, fmt) in enumerate(info_items):
        r = row + 1 + i
        _write_kv(ws, r, 1, label, 2, val, fmt)

    # --- Market Data ---
    row = 4
    mkt_items = [
        ("Current Price", stock["current_price"], FMT_PRICE),
        ("Market Cap", stock["market_cap"], FMT_DOLLAR_B),
        ("Shares Outstanding", stock["shares_outstanding"], FMT_NUM),
        ("Beta", stock["beta"], "0.00"),
        ("Trailing P/E", stock["trailing_pe"], "0.0x"),
        ("Dividend Yield", stock["dividend_yield"], FMT_PCT),
    ]
    for i, (label, val, fmt) in enumerate(mkt_items):
        r = row + 1 + i
        _write_kv(ws, r, 4, label, 5, val, fmt)

    # --- Interest Rates ---
    rate_items = [
        ("10-Year Treasury", rates["treasury_10y"], FMT_PCT2),
        ("2-Year Treasury", rates["treasury_2y"], FMT_PCT2),
        ("Fed Funds Rate", rates["fed_funds_rate"], FMT_PCT2),
        ("Rate Data Date", rates["date_fetched"], None),
        ("WACC", model_result["wacc_data"]["wacc"], FMT_PCT2),
        ("Cost of Equity", model_result["wacc_data"]["cost_of_equity"], FMT_PCT2),
    ]
    for i, (label, val, fmt) in enumerate(rate_items):
        r = row + 1 + i
        _write_kv(ws, r, 7, label, 8, val, fmt)

    # --- Scenario Valuation Summary ---
    row = 12
    _write_section_header(ws, row, "SCENARIO VALUATION SUMMARY", max_col)

    row = 13
    headers = ["Scenario", "Description", "", "WACC", "Terminal Growth",
               "Enterprise Value", "Equity Value", "Implied Price",
               "Current Price", "Upside/Downside"]
    _write_header_row(ws, row, headers)

    scenario_order = ["bull", "base", "bear", "rate_hike", "rate_cut"]
    for i, key in enumerate(scenario_order):
        r = row + 1 + i
        dcf = model_result["scenarios"][key]["dcf"]
        alt = i % 2 == 1

        vals = [
            dcf["scenario"],
            dcf["scenario_description"],
            "",
            dcf["wacc"],
            dcf["terminal_growth"],
            dcf["enterprise_value"],
            dcf["equity_value"],
            dcf["implied_share_price"],
            dcf["current_price"],
            dcf["upside_downside"],
        ]
        fmts = [None, None, None, FMT_PCT2, FMT_PCT2,
                FMT_DOLLAR_B, FMT_DOLLAR_B, FMT_PRICE, FMT_PRICE, FMT_PCT]

        for j, (v, f) in enumerate(zip(vals, fmts)):
            cell = ws.cell(row=r, column=1 + j, value=v)
            cell.font = VALUE_FONT
            cell.alignment = RIGHT if j >= 3 else LEFT
            cell.border = THIN_BORDER
            if f:
                cell.number_format = f
            if alt:
                cell.fill = ALT_ROW_FILL

        # Color the upside/downside
        ud_cell = ws.cell(row=r, column=10)
        if dcf["upside_downside"] > 0:
            ud_cell.font = GREEN_FONT
        else:
            ud_cell.font = RED_FONT

    # --- Historical Summary ---
    row = 20
    _write_section_header(ws, row, "HISTORICAL FINANCIAL SUMMARY (in Billions)", max_col)

    row = 21
    year_headers = ["Metric"] + financials["years"] + [""]*(max_col - 1 - len(financials["years"]))
    _write_header_row(ws, row, year_headers[:max_col])

    hist_items = [
        ("Revenue", financials["revenue"]),
        ("Gross Profit", financials["gross_profit"]),
        ("Operating Income", financials["operating_income"]),
        ("Net Income", financials["net_income"]),
        ("Free Cash Flow", financials["free_cash_flow"]),
        ("Total Debt", financials["total_debt"]),
        ("Cash & Equivalents", financials["cash"]),
        ("Total Equity", financials["total_equity"]),
    ]
    for i, (label, vals) in enumerate(hist_items):
        r = row + 1 + i
        _write_data_row(ws, r, label, vals, fmt=FMT_DOLLAR_B, alt=i % 2 == 1)

    # --- Instruction note ---
    row = 31
    ws.merge_cells(start_row=row, start_column=1, end_row=row + 1, end_column=max_col)
    note = ws.cell(row=row, column=1,
                   value="HOW TO USE: Run  python generate_dcf.py <TICKER>  to regenerate this workbook for any stock. "
                         "Example: python generate_dcf.py MSFT. Requires yfinance and openpyxl (pip install -r requirements.txt).")
    note.font = Font(name="Calibri", size=10, italic=True, color="666666")
    note.alignment = WRAP

    # ===== CHARTS =====
    years = financials["years"]
    n_years = len(years)
    # Chart data area starts below the visible content
    cd = 50  # chart data start row

    # -- Chart 1: Revenue vs FCF bar chart --
    ws.cell(row=cd, column=1, value="Year")
    ws.cell(row=cd + 1, column=1, value="Revenue ($B)")
    ws.cell(row=cd + 2, column=1, value="Free Cash Flow ($B)")
    ws.cell(row=cd + 3, column=1, value="Net Income ($B)")
    for i, yr in enumerate(years):
        ws.cell(row=cd, column=2 + i, value=yr)
        ws.cell(row=cd + 1, column=2 + i, value=financials["revenue"][i] / 1e9)
        ws.cell(row=cd + 2, column=2 + i, value=financials["free_cash_flow"][i] / 1e9)
        ws.cell(row=cd + 3, column=2 + i, value=financials["net_income"][i] / 1e9)

    _make_bar_chart(ws, "Revenue vs FCF vs Net Income ($B)", "$ Billions",
                    [cd + 1, cd + 2, cd + 3], cd, 2, 1 + n_years,
                    ["Revenue", "Free Cash Flow", "Net Income"],
                    "A33", width=28, height=14)

    # -- Chart 2: Profitability margins line chart --
    ws.cell(row=cd + 5, column=1, value="Year")
    ws.cell(row=cd + 6, column=1, value="Gross Margin")
    ws.cell(row=cd + 7, column=1, value="Operating Margin")
    ws.cell(row=cd + 8, column=1, value="Net Margin")
    ws.cell(row=cd + 9, column=1, value="FCF Margin")
    for i, yr in enumerate(years):
        rev = financials["revenue"][i] or 1
        ws.cell(row=cd + 5, column=2 + i, value=yr)
        ws.cell(row=cd + 6, column=2 + i, value=financials["gross_profit"][i] / rev)
        ws.cell(row=cd + 7, column=2 + i, value=financials["operating_income"][i] / rev)
        ws.cell(row=cd + 8, column=2 + i, value=financials["net_income"][i] / rev)
        ws.cell(row=cd + 9, column=2 + i, value=financials["free_cash_flow"][i] / rev)

    chart2 = _make_line_chart(ws, "Profitability Margins Over Time", "Margin %",
                              [cd + 6, cd + 7, cd + 8, cd + 9], cd + 5, 2, 1 + n_years,
                              ["Gross Margin", "Operating Margin", "Net Margin", "FCF Margin"],
                              "F33", width=24, height=14)
    chart2.y_axis.numFmt = '0%'

    # -- Chart 3: Scenario implied price bar chart --
    ws.cell(row=cd + 11, column=1, value="Scenario")
    ws.cell(row=cd + 12, column=1, value="Implied Price ($)")
    ws.cell(row=cd + 13, column=1, value="Current Price ($)")
    for i, key in enumerate(scenario_order):
        dcf = model_result["scenarios"][key]["dcf"]
        ws.cell(row=cd + 11, column=2 + i, value=dcf["scenario"])
        ws.cell(row=cd + 12, column=2 + i, value=dcf["implied_share_price"])
        ws.cell(row=cd + 13, column=2 + i, value=dcf["current_price"])

    _make_bar_chart(ws, "Implied Share Price by Scenario", "Price ($)",
                    [cd + 12, cd + 13], cd + 11, 2, 1 + len(scenario_order),
                    ["Implied Price", "Current Price"],
                    "A49", width=28, height=14)

    # -- Chart 4: WACC components stacked view --
    ws.cell(row=cd + 15, column=1, value="Scenario")
    ws.cell(row=cd + 16, column=1, value="WACC (%)")
    ws.cell(row=cd + 17, column=1, value="Terminal Growth (%)")
    for i, key in enumerate(scenario_order):
        dcf = model_result["scenarios"][key]["dcf"]
        ws.cell(row=cd + 15, column=2 + i, value=dcf["scenario"])
        ws.cell(row=cd + 16, column=2 + i, value=dcf["wacc"])
        ws.cell(row=cd + 17, column=2 + i, value=dcf["terminal_growth"])

    chart4 = _make_bar_chart(ws, "WACC vs Terminal Growth by Scenario", "Rate",
                             [cd + 16, cd + 17], cd + 15, 2, 1 + len(scenario_order),
                             ["WACC", "Terminal Growth"],
                             "F49", width=24, height=14)
    chart4.y_axis.numFmt = '0.0%'


def build_financial_statement_sheet(wb, sheet_name, financials, items, tab_color,
                                    chart_config=None):
    """Generic builder for income statement / balance sheet / cash flow sheets."""
    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_properties.tabColor = tab_color
    years = financials["years"]
    max_col = 1 + len(years)

    _set_col_widths(ws, {1: 32})
    for i in range(2, max_col + 1):
        ws.column_dimensions[get_column_letter(i)].width = 18

    row = 1
    _write_title_row(ws, row, f"  {sheet_name.upper()}", max_col)

    row = 2
    _write_header_row(ws, row, ["Line Item"] + years)

    for i, (label, key, fmt, is_section) in enumerate(items):
        row += 1
        if is_section:
            _write_section_header(ws, row, label, max_col)
            continue

        vals = financials.get(key, [0] * len(years))
        alt = (row % 2 == 0)
        bold = label.startswith("=")
        if bold:
            label = label[1:]

        _write_data_row(ws, row, label, vals, fmt=fmt, alt=alt,
                        label_font=BOLD_VALUE if bold else LABEL_FONT,
                        value_font=BOLD_VALUE if bold else VALUE_FONT)

    # Add chart if config provided
    if chart_config:
        _add_statement_chart(ws, financials, row + 2, chart_config)

    return ws


def _add_statement_chart(ws, financials, start_row, config):
    """Add a chart to a financial statement sheet."""
    years = financials["years"]
    n = len(years)
    cd = start_row + 16  # data area below chart

    ws.cell(row=cd, column=1, value="Year")
    for i, yr in enumerate(years):
        ws.cell(row=cd, column=2 + i, value=yr)

    data_rows = []
    labels = []
    for offset, (key, label, divisor) in enumerate(config["series"]):
        r = cd + 1 + offset
        ws.cell(row=r, column=1, value=label)
        vals = financials.get(key, [0] * n)
        for i, v in enumerate(vals):
            ws.cell(row=r, column=2 + i, value=(v or 0) / divisor)
        data_rows.append(r)
        labels.append(label)

    chart_type = config.get("chart_type", "bar")
    if chart_type == "line":
        chart = _make_line_chart(ws, config["title"], config["y_title"],
                                 data_rows, cd, 2, 1 + n, labels,
                                 f"A{start_row}", width=28, height=14)
    else:
        chart = _make_bar_chart(ws, config["title"], config["y_title"],
                                data_rows, cd, 2, 1 + n, labels,
                                f"A{start_row}", width=28, height=14)

    if config.get("pct_format"):
        chart.y_axis.numFmt = '0%'

    return chart


def build_income_statement(wb, financials):
    items = [
        ("INCOME STATEMENT", None, None, True),
        ("Revenue", "revenue", FMT_DOLLAR_B, False),
        ("Cost of Revenue", "cost_of_revenue", FMT_DOLLAR_B, False),
        ("=Gross Profit", "gross_profit", FMT_DOLLAR_B, False),
        ("Operating Income (EBIT)", "operating_income", FMT_DOLLAR_B, False),
        ("EBITDA", "ebitda", FMT_DOLLAR_B, False),
        ("Interest Expense", "interest_expense", FMT_DOLLAR_B, False),
        ("Tax Provision", "tax_provision", FMT_DOLLAR_B, False),
        ("=Net Income", "net_income", FMT_DOLLAR_B, False),
        ("Depreciation & Amortization", "depreciation", FMT_DOLLAR_B, False),
    ]
    chart_cfg = {
        "title": "Revenue, Gross Profit & Net Income ($B)",
        "y_title": "$ Billions",
        "series": [
            ("revenue", "Revenue ($B)", 1e9),
            ("gross_profit", "Gross Profit ($B)", 1e9),
            ("operating_income", "Operating Income ($B)", 1e9),
            ("net_income", "Net Income ($B)", 1e9),
        ],
    }
    ws = build_financial_statement_sheet(wb, "Income Statement", financials, items,
                                         MED_BLUE, chart_cfg)

    # Add margin chart below
    years = financials["years"]
    n = len(years)
    cd2 = 40
    ws.cell(row=cd2, column=1, value="Year")
    ws.cell(row=cd2 + 1, column=1, value="Gross Margin")
    ws.cell(row=cd2 + 2, column=1, value="Operating Margin")
    ws.cell(row=cd2 + 3, column=1, value="Net Margin")
    for i, yr in enumerate(years):
        rev = financials["revenue"][i] or 1
        ws.cell(row=cd2, column=2 + i, value=yr)
        ws.cell(row=cd2 + 1, column=2 + i, value=financials["gross_profit"][i] / rev)
        ws.cell(row=cd2 + 2, column=2 + i, value=financials["operating_income"][i] / rev)
        ws.cell(row=cd2 + 3, column=2 + i, value=financials["net_income"][i] / rev)

    mc = _make_line_chart(ws, "Profit Margins Over Time", "Margin",
                          [cd2 + 1, cd2 + 2, cd2 + 3], cd2, 2, 1 + n,
                          ["Gross Margin", "Operating Margin", "Net Margin"],
                          "D14", width=24, height=14)
    mc.y_axis.numFmt = '0%'


def build_balance_sheet(wb, financials):
    items = [
        ("ASSETS", None, None, True),
        ("Current Assets", "current_assets", FMT_DOLLAR_B, False),
        ("Cash & Equivalents", "cash", FMT_DOLLAR_B, False),
        ("=Total Assets", "total_assets", FMT_DOLLAR_B, False),
        ("LIABILITIES & EQUITY", None, None, True),
        ("Current Liabilities", "current_liabilities", FMT_DOLLAR_B, False),
        ("Total Debt", "total_debt", FMT_DOLLAR_B, False),
        ("=Total Liabilities", "total_liabilities", FMT_DOLLAR_B, False),
        ("=Total Stockholders' Equity", "total_equity", FMT_DOLLAR_B, False),
    ]
    chart_cfg = {
        "title": "Balance Sheet Structure ($B)",
        "y_title": "$ Billions",
        "series": [
            ("total_assets", "Total Assets ($B)", 1e9),
            ("total_liabilities", "Total Liabilities ($B)", 1e9),
            ("total_equity", "Total Equity ($B)", 1e9),
            ("total_debt", "Total Debt ($B)", 1e9),
        ],
    }
    ws = build_financial_statement_sheet(wb, "Balance Sheet", financials, items,
                                         "2E7D32", chart_cfg)

    # Add Debt-to-Equity ratio line chart
    years = financials["years"]
    n = len(years)
    cd2 = 42
    ws.cell(row=cd2, column=1, value="Year")
    ws.cell(row=cd2 + 1, column=1, value="Debt-to-Equity Ratio")
    ws.cell(row=cd2 + 2, column=1, value="Current Ratio")
    for i, yr in enumerate(years):
        eq = financials["total_equity"][i] or 1
        cl = financials["current_liabilities"][i] or 1
        ws.cell(row=cd2, column=2 + i, value=yr)
        ws.cell(row=cd2 + 1, column=2 + i, value=financials["total_debt"][i] / eq)
        ws.cell(row=cd2 + 2, column=2 + i, value=financials["current_assets"][i] / cl)

    rc = _make_line_chart(ws, "Key Ratios Over Time", "Ratio",
                          [cd2 + 1, cd2 + 2], cd2, 2, 1 + n,
                          ["Debt-to-Equity", "Current Ratio"],
                          "D13", width=24, height=14)
    rc.y_axis.numFmt = '0.00x'


def build_cash_flow(wb, financials):
    items = [
        ("CASH FLOW STATEMENT", None, None, True),
        ("Operating Cash Flow", "operating_cash_flow", FMT_DOLLAR_B, False),
        ("Capital Expenditure", "capex", FMT_DOLLAR_B, False),
        ("Depreciation & Amortization", "depreciation_amortization", FMT_DOLLAR_B, False),
        ("Change in Working Capital", "change_in_working_capital", FMT_DOLLAR_B, False),
        ("=Free Cash Flow (OCF + CapEx)", "free_cash_flow", FMT_DOLLAR_B, False),
    ]
    chart_cfg = {
        "title": "Cash Flow Components ($B)",
        "y_title": "$ Billions",
        "series": [
            ("operating_cash_flow", "Operating CF ($B)", 1e9),
            ("free_cash_flow", "Free Cash Flow ($B)", 1e9),
        ],
    }
    ws = build_financial_statement_sheet(wb, "Cash Flow", financials, items,
                                         ACCENT_ORANGE, chart_cfg)

    # Add FCF conversion chart
    years = financials["years"]
    n = len(years)
    cd2 = 36
    ws.cell(row=cd2, column=1, value="Year")
    ws.cell(row=cd2 + 1, column=1, value="FCF Yield (FCF/Revenue)")
    ws.cell(row=cd2 + 2, column=1, value="CapEx % of Revenue")
    for i, yr in enumerate(years):
        rev = financials["revenue"][i] or 1
        ws.cell(row=cd2, column=2 + i, value=yr)
        ws.cell(row=cd2 + 1, column=2 + i, value=financials["free_cash_flow"][i] / rev)
        ws.cell(row=cd2 + 2, column=2 + i, value=abs(financials["capex"][i]) / rev)

    fc = _make_line_chart(ws, "FCF Yield & CapEx Intensity", "% of Revenue",
                          [cd2 + 1, cd2 + 2], cd2, 2, 1 + n,
                          ["FCF Yield", "CapEx Intensity"],
                          "D10", width=24, height=14)
    fc.y_axis.numFmt = '0%'


def build_wacc_sheet(wb, wacc_data, stock, financials, rates):
    """Sheet 5: WACC breakdown with charts."""
    ws = wb.create_sheet(title="WACC")
    ws.sheet_properties.tabColor = "8E24AA"  # purple
    max_col = 6

    _set_col_widths(ws, {1: 32, 2: 18, 3: 5, 4: 32, 5: 18, 6: 5})

    row = 1
    _write_title_row(ws, row, "  WEIGHTED AVERAGE COST OF CAPITAL (WACC)", max_col)

    # WACC Formula
    row = 3
    _write_section_header(ws, row, "WACC = (E/V) x Re  +  (D/V) x Rd x (1 - T)", max_col)

    # Cost of Equity (CAPM)
    row = 5
    _write_section_header(ws, row, "COST OF EQUITY  (CAPM: Re = Rf + B x ERP)", max_col)

    capm_items = [
        ("Risk-Free Rate (Rf)", wacc_data["risk_free_rate"], FMT_PCT2),
        ("Beta (B)", wacc_data["beta"], "0.00"),
        ("Equity Risk Premium (ERP)", wacc_data["equity_risk_premium"], FMT_PCT2),
        ("Cost of Equity (Re)", wacc_data["cost_of_equity"], FMT_PCT2),
    ]
    for i, (label, val, fmt) in enumerate(capm_items):
        r = row + 1 + i
        _write_kv(ws, r, 1, label, 2, val, fmt,
                  value_font=BOLD_VALUE if "Cost of Equity" in label else VALUE_FONT)

    # Cost of Debt
    row = 11
    _write_section_header(ws, row, "COST OF DEBT  (Rd = Interest Expense / Total Debt)", max_col)

    debt_items = [
        ("Interest Expense", abs(financials["interest_expense"][0]), FMT_DOLLAR_B),
        ("Total Debt", financials["total_debt"][0], FMT_DOLLAR_B),
        ("Cost of Debt (Rd)", wacc_data["cost_of_debt"], FMT_PCT2),
        ("Effective Tax Rate (T)", wacc_data["tax_rate"], FMT_PCT2),
        ("After-Tax Cost of Debt", wacc_data["cost_of_debt"] * (1 - wacc_data["tax_rate"]), FMT_PCT2),
    ]
    for i, (label, val, fmt) in enumerate(debt_items):
        r = row + 1 + i
        _write_kv(ws, r, 1, label, 2, val, fmt)

    # Capital Structure
    row = 18
    _write_section_header(ws, row, "CAPITAL STRUCTURE", max_col)

    cap_items = [
        ("Market Cap (Equity Value)", wacc_data["equity_value"], FMT_DOLLAR_B),
        ("Total Debt (Debt Value)", wacc_data["debt_value"], FMT_DOLLAR_B),
        ("Total Capital (V = E + D)", wacc_data["equity_value"] + wacc_data["debt_value"], FMT_DOLLAR_B),
        ("Weight of Equity (E/V)", wacc_data["weight_equity"], FMT_PCT2),
        ("Weight of Debt (D/V)", wacc_data["weight_debt"], FMT_PCT2),
    ]
    for i, (label, val, fmt) in enumerate(cap_items):
        r = row + 1 + i
        _write_kv(ws, r, 1, label, 2, val, fmt)

    # Final WACC
    row = 25
    _write_section_header(ws, row, "RESULT", max_col)
    row = 26
    _write_kv(ws, row, 1, "WACC", 2, wacc_data["wacc"], FMT_PCT2,
              label_font=Font(name="Calibri", size=14, bold=True, color=DARK_BLUE),
              value_font=BIG_NUMBER)

    # Interest Rate Environment
    row = 28
    _write_section_header(ws, row, "CURRENT INTEREST RATE ENVIRONMENT", max_col)
    rate_items = [
        ("10-Year Treasury Yield", rates["treasury_10y"], FMT_PCT2),
        ("2-Year Treasury Yield", rates["treasury_2y"], FMT_PCT2),
        ("Federal Funds Rate", rates["fed_funds_rate"], FMT_PCT2),
        ("Yield Curve Spread (10Y - 2Y)", rates["treasury_10y"] - rates["treasury_2y"], FMT_PCT2),
        ("Data Retrieved", rates["date_fetched"], None),
    ]
    for i, (label, val, fmt) in enumerate(rate_items):
        r = row + 1 + i
        _write_kv(ws, r, 1, label, 2, val, fmt)

    # ===== CHARTS =====

    # -- Chart 1: Capital Structure Pie Chart --
    cd = 45
    ws.cell(row=cd, column=1, value="Component")
    ws.cell(row=cd + 1, column=1, value="Equity")
    ws.cell(row=cd + 2, column=1, value="Debt")
    ws.cell(row=cd, column=2, value="Value ($B)")
    ws.cell(row=cd + 1, column=2, value=wacc_data["equity_value"] / 1e9)
    ws.cell(row=cd + 2, column=2, value=wacc_data["debt_value"] / 1e9)

    pie = PieChart()
    pie.title = "Capital Structure (Equity vs Debt)"
    pie.style = 10
    pie.width = 16
    pie.height = 14
    data_ref = Reference(ws, min_col=2, max_col=2, min_row=cd, max_row=cd + 2)
    cats_ref = Reference(ws, min_col=1, max_col=1, min_row=cd + 1, max_row=cd + 2)
    pie.add_data(data_ref, titles_from_data=True)
    pie.set_categories(cats_ref)
    # Color the pie slices
    s = pie.series[0]
    pt0 = DataPoint(idx=0)
    pt0.graphicalProperties.solidFill = MED_BLUE
    pt1 = DataPoint(idx=1)
    pt1.graphicalProperties.solidFill = ACCENT_ORANGE
    s.data_points = [pt0, pt1]
    s.dLbls = DataLabelList()
    s.dLbls.showPercent = True
    ws.add_chart(pie, "D5")

    # -- Chart 2: WACC Components Bar Chart --
    ws.cell(row=cd + 4, column=1, value="Component")
    ws.cell(row=cd + 5, column=1, value="Rate (%)")
    components = [
        ("Risk-Free Rate", wacc_data["risk_free_rate"]),
        ("Cost of Equity", wacc_data["cost_of_equity"]),
        ("Cost of Debt", wacc_data["cost_of_debt"]),
        ("After-Tax CoD", wacc_data["cost_of_debt"] * (1 - wacc_data["tax_rate"])),
        ("WACC", wacc_data["wacc"]),
    ]
    for j, (name, val) in enumerate(components):
        ws.cell(row=cd + 4, column=2 + j, value=name)
        ws.cell(row=cd + 5, column=2 + j, value=val)

    chart_w = BarChart()
    chart_w.type = "col"
    chart_w.style = 10
    chart_w.title = "WACC Component Rates"
    chart_w.y_axis.title = "Rate"
    chart_w.y_axis.numFmt = '0.0%'
    _style_chart(chart_w, 24, 14)

    cats_w = Reference(ws, min_col=2, max_col=1 + len(components), min_row=cd + 4)
    data_w = Reference(ws, min_col=2, max_col=1 + len(components), min_row=cd + 5)
    chart_w.add_data(data_w, from_rows=True, titles_from_data=False)
    chart_w.set_categories(cats_w)
    chart_w.series[0].tx = SeriesLabel(v="Rate")
    _color_series(chart_w)
    ws.add_chart(chart_w, "D19")

    # -- Chart 3: Interest Rate Environment Bar --
    ws.cell(row=cd + 7, column=1, value="Rate")
    ws.cell(row=cd + 8, column=1, value="Yield (%)")
    rate_bars = [
        ("Fed Funds", rates["fed_funds_rate"]),
        ("2Y Treasury", rates["treasury_2y"]),
        ("10Y Treasury", rates["treasury_10y"]),
        ("WACC", wacc_data["wacc"]),
    ]
    for j, (name, val) in enumerate(rate_bars):
        ws.cell(row=cd + 7, column=2 + j, value=name)
        ws.cell(row=cd + 8, column=2 + j, value=val)

    chart_r = BarChart()
    chart_r.type = "col"
    chart_r.style = 10
    chart_r.title = "Rate Environment: Treasury Yields vs WACC"
    chart_r.y_axis.title = "Rate"
    chart_r.y_axis.numFmt = '0.0%'
    _style_chart(chart_r, 24, 14)

    cats_r = Reference(ws, min_col=2, max_col=1 + len(rate_bars), min_row=cd + 7)
    data_r = Reference(ws, min_col=2, max_col=1 + len(rate_bars), min_row=cd + 8)
    chart_r.add_data(data_r, from_rows=True, titles_from_data=False)
    chart_r.set_categories(cats_r)
    chart_r.series[0].tx = SeriesLabel(v="Yield")
    _color_series(chart_r)
    ws.add_chart(chart_r, "A35")


def build_dcf_scenario_sheet(wb, scenario_key, model_result, financials, stock):
    """Build a detailed DCF sheet for one scenario with projection charts."""
    sc_data = model_result["scenarios"][scenario_key]
    scenario = sc_data["scenario"]
    proj = sc_data["projection"]
    dcf = sc_data["dcf"]
    wacc_data = model_result["wacc_data"]

    tab_colors = {
        "base": MED_BLUE, "bull": ACCENT_GREEN, "bear": ACCENT_RED,
        "rate_hike": ACCENT_ORANGE, "rate_cut": ACCENT_TEAL,
    }
    sheet_names = {
        "base": "DCF Base Case", "bull": "DCF Bull Case", "bear": "DCF Bear Case",
        "rate_hike": "DCF Rising Rates", "rate_cut": "DCF Falling Rates",
    }

    ws = wb.create_sheet(title=sheet_names[scenario_key])
    ws.sheet_properties.tabColor = tab_colors.get(scenario_key, MED_BLUE)

    n_proj = proj["projection_years"]
    hist_years = financials["years"]
    proj_years = [str(int(hist_years[0]) + i + 1) for i in range(n_proj)]
    all_years = hist_years + [""] + proj_years + ["Terminal"]
    max_col = 1 + len(all_years)

    _set_col_widths(ws, {1: 30})
    for i in range(2, max_col + 1):
        ws.column_dimensions[get_column_letter(i)].width = 16

    # Title
    row = 1
    _write_title_row(ws, row, f"  DCF MODEL  —  {scenario.name.upper()}", max_col)
    row = 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row=row, column=1, value=f"  {scenario.description}")
    cell.font = Font(name="Calibri", size=10, italic=True, color=LIGHT_BLUE)
    cell.fill = TITLE_FILL
    ws.row_dimensions[row].height = 22

    # Scenario adjustments
    row = 4
    _write_section_header(ws, row, "SCENARIO ASSUMPTIONS", max_col)
    adj_items = [
        ("Revenue Growth Adjustment", scenario.revenue_growth_adj, FMT_PCT),
        ("Operating Margin Adjustment", scenario.margin_adj, FMT_PCT),
        ("WACC Adjustment", scenario.wacc_adj, FMT_PCT),
        ("Terminal Growth Adjustment", scenario.terminal_growth_adj, FMT_PCT),
        ("Scenario WACC", dcf["wacc"], FMT_PCT2),
        ("Terminal Growth Rate", dcf["terminal_growth"], FMT_PCT2),
    ]
    for i, (label, val, fmt) in enumerate(adj_items):
        r = row + 1 + i
        _write_kv(ws, r, 1, label, 2, val, fmt)

    # Historical + Projected FCF
    row = 12
    _write_section_header(ws, row, "HISTORICAL  <-->  PROJECTED", max_col)

    row = 13
    header_labels = [""] + hist_years + ["->"] + proj_years + ["Terminal"]
    _write_header_row(ws, row, header_labels)

    # Revenue
    row = 14
    hist_rev = financials["revenue"]
    proj_rev = proj["projected_revenue"]
    term_rev = proj_rev[-1] * (1 + dcf["terminal_growth"])
    _write_data_row(ws, row, "Revenue", hist_rev + [None] + proj_rev + [term_rev], fmt=FMT_DOLLAR_B)

    # Growth Rate
    row = 15
    hist_growths = model_result["metrics"]["revenue_growths"]
    proj_growths = proj["growth_rates"]
    growth_vals = hist_growths + [None] * (len(hist_years) - len(hist_growths)) + [None] + proj_growths + [dcf["terminal_growth"]]
    _write_data_row(ws, row, "Revenue Growth %", growth_vals, fmt=FMT_PCT, alt=True)

    # Operating Income / EBIT
    row = 16
    hist_ebit = financials["operating_income"]
    proj_ebit = proj["projected_ebit"]
    term_ebit = proj_ebit[-1] * (1 + dcf["terminal_growth"])
    _write_data_row(ws, row, "EBIT (Operating Income)", hist_ebit + [None] + proj_ebit + [term_ebit], fmt=FMT_DOLLAR_B)

    # Operating Margin
    row = 17
    hist_margins = model_result["metrics"]["operating_margins"]
    proj_margins = proj["margins"]
    margin_vals = hist_margins + [None] * (len(hist_years) - len(hist_margins)) + [None] + proj_margins + [proj_margins[-1]]
    _write_data_row(ws, row, "Operating Margin %", margin_vals, fmt=FMT_PCT, alt=True)

    # NOPAT
    row = 18
    hist_nopat = [financials["operating_income"][i] * (1 - proj["tax_rate"]) for i in range(len(hist_years))]
    proj_nopat = proj["projected_nopat"]
    term_nopat = proj_nopat[-1] * (1 + dcf["terminal_growth"])
    _write_data_row(ws, row, "NOPAT (EBIT x (1-T))", hist_nopat + [None] + proj_nopat + [term_nopat], fmt=FMT_DOLLAR_B)

    # D&A
    row = 19
    hist_da = financials["depreciation"]
    proj_da = proj["projected_da"]
    term_da = proj_da[-1] * (1 + dcf["terminal_growth"])
    _write_data_row(ws, row, "(+) Depreciation & Amort.", hist_da + [None] + proj_da + [term_da], fmt=FMT_DOLLAR_B, alt=True)

    # CapEx
    row = 20
    hist_capex = [abs(c) for c in financials["capex"]]
    proj_capex = proj["projected_capex"]
    term_capex = proj_capex[-1] * (1 + dcf["terminal_growth"])
    _write_data_row(ws, row, "(-) Capital Expenditure", hist_capex + [None] + proj_capex + [term_capex], fmt=FMT_DOLLAR_B)

    # FCF
    row = 21
    hist_fcf = financials["free_cash_flow"]
    proj_fcf = proj["projected_fcf"]
    term_fcf = proj_fcf[-1] * (1 + dcf["terminal_growth"])
    _write_data_row(ws, row, "=Unlevered Free Cash Flow",
                    hist_fcf + [None] + proj_fcf + [term_fcf],
                    fmt=FMT_DOLLAR_B, label_font=BOLD_VALUE, value_font=BOLD_VALUE, alt=True)

    # PV of FCFs
    row = 23
    _write_section_header(ws, row, "PRESENT VALUE CALCULATION", max_col)

    row = 24
    pv_vals = [None] * len(hist_years) + [None] + dcf["pv_fcfs"] + [dcf["pv_terminal_value"]]
    _write_data_row(ws, row, "PV of Free Cash Flow", pv_vals, fmt=FMT_DOLLAR_B)

    row = 25
    disc_factors = [None] * len(hist_years) + [None] + \
                   [1 / (1 + dcf["wacc"]) ** (i + 1) for i in range(n_proj)] + \
                   [1 / (1 + dcf["wacc"]) ** n_proj]
    _write_data_row(ws, row, "Discount Factor", disc_factors, fmt="0.0000", alt=True)

    # Valuation Bridge
    row = 27
    _write_section_header(ws, row, "VALUATION BRIDGE", max_col)

    bridge_items = [
        ("Sum of PV(FCFs)", dcf["pv_fcf_total"], FMT_DOLLAR_B),
        ("Terminal Value (undiscounted)", dcf["terminal_value"], FMT_DOLLAR_B),
        ("PV of Terminal Value", dcf["pv_terminal_value"], FMT_DOLLAR_B),
        ("= Enterprise Value", dcf["enterprise_value"], FMT_DOLLAR_B),
        ("(-) Net Debt", dcf["net_debt"], FMT_DOLLAR_B),
        ("= Equity Value", dcf["equity_value"], FMT_DOLLAR_B),
        ("Shares Outstanding", dcf["shares_outstanding"], FMT_NUM),
        ("= Implied Share Price", dcf["implied_share_price"], FMT_PRICE),
        ("Current Market Price", dcf["current_price"], FMT_PRICE),
        ("Upside / Downside", dcf["upside_downside"], FMT_PCT),
    ]
    for i, (label, val, fmt) in enumerate(bridge_items):
        r = row + 1 + i
        bold = label.startswith("=")
        lbl = label.lstrip("= ")
        _write_kv(ws, r, 1, lbl, 2, val, fmt,
                  label_font=BOLD_VALUE if bold else LABEL_FONT,
                  value_font=BOLD_VALUE if bold else VALUE_FONT)

    # Color implied price
    price_cell = ws.cell(row=row + 8, column=2)
    if dcf["upside_downside"] > 0:
        price_cell.font = Font(name="Calibri", size=14, bold=True, color=ACCENT_GREEN)
    else:
        price_cell.font = Font(name="Calibri", size=14, bold=True, color=ACCENT_RED)

    ud_cell = ws.cell(row=row + 10, column=2)
    if dcf["upside_downside"] > 0:
        ud_cell.font = GREEN_FONT
        ud_cell.fill = GREEN_FILL
    else:
        ud_cell.font = RED_FONT
        ud_cell.fill = RED_FILL

    # ===== CHARTS =====
    combined_years = hist_years + proj_years
    n_all = len(combined_years)
    cd = 50

    # -- Chart 1: Revenue Projection (Historical + Projected) --
    ws.cell(row=cd, column=1, value="Year")
    ws.cell(row=cd + 1, column=1, value="Historical Revenue ($B)")
    ws.cell(row=cd + 2, column=1, value="Projected Revenue ($B)")
    for i, yr in enumerate(hist_years):
        ws.cell(row=cd, column=2 + i, value=yr)
        ws.cell(row=cd + 1, column=2 + i, value=hist_rev[i] / 1e9)
        ws.cell(row=cd + 2, column=2 + i, value=None)  # blank projected for hist
    for i, yr in enumerate(proj_years):
        col = 2 + len(hist_years) + i
        ws.cell(row=cd, column=col, value=yr)
        ws.cell(row=cd + 1, column=col, value=None)  # blank historical for proj
        ws.cell(row=cd + 2, column=col, value=proj_rev[i] / 1e9)

    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = f"Revenue Projection — {scenario.name} ($B)"
    chart1.y_axis.title = "$ Billions"
    _style_chart(chart1, 28, 14)
    cats1 = Reference(ws, min_col=2, max_col=1 + n_all, min_row=cd)
    d1a = Reference(ws, min_col=2, max_col=1 + n_all, min_row=cd + 1)
    d1b = Reference(ws, min_col=2, max_col=1 + n_all, min_row=cd + 2)
    chart1.add_data(d1a, from_rows=True, titles_from_data=False)
    chart1.add_data(d1b, from_rows=True, titles_from_data=False)
    chart1.set_categories(cats1)
    chart1.series[0].tx = SeriesLabel(v="Historical")
    chart1.series[1].tx = SeriesLabel(v="Projected")
    chart1.series[0].graphicalProperties.solidFill = MED_BLUE
    chart1.series[1].graphicalProperties.solidFill = ACCENT_GREEN if scenario_key in ["bull", "rate_cut"] else (ACCENT_RED if scenario_key in ["bear", "rate_hike"] else ACCENT_ORANGE)
    ws.add_chart(chart1, "D4")

    # -- Chart 2: FCF Projection with PV overlay --
    ws.cell(row=cd + 4, column=1, value="Year")
    ws.cell(row=cd + 5, column=1, value="Projected FCF ($B)")
    ws.cell(row=cd + 6, column=1, value="PV of FCF ($B)")
    for i, yr in enumerate(proj_years):
        ws.cell(row=cd + 4, column=2 + i, value=yr)
        ws.cell(row=cd + 5, column=2 + i, value=proj_fcf[i] / 1e9)
        ws.cell(row=cd + 6, column=2 + i, value=dcf["pv_fcfs"][i] / 1e9)

    _make_bar_chart(ws, f"FCF vs Present Value — {scenario.name} ($B)", "$ Billions",
                    [cd + 5, cd + 6], cd + 4, 2, 1 + n_proj,
                    ["Projected FCF", "PV of FCF"],
                    "D27", width=28, height=14)

    # -- Chart 3: Valuation Bridge Waterfall --
    ws.cell(row=cd + 8, column=1, value="Step")
    ws.cell(row=cd + 9, column=1, value="Value ($B)")
    bridge_chart_data = [
        ("PV of FCFs", dcf["pv_fcf_total"] / 1e9),
        ("PV Terminal", dcf["pv_terminal_value"] / 1e9),
        ("Enterprise V.", dcf["enterprise_value"] / 1e9),
        ("(-) Net Debt", -dcf["net_debt"] / 1e9),
        ("Equity Value", dcf["equity_value"] / 1e9),
    ]
    for j, (label, val) in enumerate(bridge_chart_data):
        ws.cell(row=cd + 8, column=2 + j, value=label)
        ws.cell(row=cd + 9, column=2 + j, value=val)

    _make_bar_chart(ws, f"Valuation Bridge — {scenario.name} ($B)", "$ Billions",
                    [cd + 9], cd + 8, 2, 1 + len(bridge_chart_data),
                    ["Value"],
                    "A39", width=28, height=14)


def build_scenario_comparison(wb, model_result, stock, financials):
    """Sheet 11: Side-by-side comparison of all scenarios with multiple charts."""
    ws = wb.create_sheet(title="Scenario Comparison")
    ws.sheet_properties.tabColor = "6A1B9A"  # deep purple
    max_col = 7

    _set_col_widths(ws, {1: 30, 2: 16, 3: 16, 4: 16, 5: 16, 6: 16, 7: 16})

    row = 1
    _write_title_row(ws, row, "  SCENARIO COMPARISON  —  ALL PATHS", max_col)

    scenario_order = ["bull", "base", "bear", "rate_hike", "rate_cut"]
    names = [model_result["scenarios"][k]["dcf"]["scenario"] for k in scenario_order]

    row = 3
    _write_header_row(ws, row, ["Metric"] + names + [""])

    comparison_rows = [
        ("WACC", "wacc", FMT_PCT2),
        ("Terminal Growth", "terminal_growth", FMT_PCT2),
        ("PV of FCFs", "pv_fcf_total", FMT_DOLLAR_B),
        ("PV of Terminal Value", "pv_terminal_value", FMT_DOLLAR_B),
        ("Enterprise Value", "enterprise_value", FMT_DOLLAR_B),
        ("Net Debt", "net_debt", FMT_DOLLAR_B),
        ("Equity Value", "equity_value", FMT_DOLLAR_B),
        ("Implied Share Price", "implied_share_price", FMT_PRICE),
        ("Current Price", "current_price", FMT_PRICE),
        ("Upside / Downside", "upside_downside", FMT_PCT),
    ]

    for i, (label, key, fmt) in enumerate(comparison_rows):
        r = row + 1 + i
        alt = i % 2 == 1
        vals = [model_result["scenarios"][k]["dcf"][key] for k in scenario_order]
        _write_data_row(ws, r, label, vals, fmt=fmt, alt=alt,
                        label_font=BOLD_VALUE if "Implied" in label or "Upside" in label else LABEL_FONT,
                        value_font=BOLD_VALUE if "Implied" in label or "Upside" in label else VALUE_FONT)

    # Color upside/downside row
    ud_row = row + len(comparison_rows)
    for j in range(len(scenario_order)):
        cell = ws.cell(row=ud_row, column=2 + j)
        if cell.value and cell.value > 0:
            cell.font = GREEN_FONT
            cell.fill = GREEN_FILL
        elif cell.value and cell.value < 0:
            cell.font = RED_FONT
            cell.fill = RED_FILL

    # Scenario Descriptions
    row = ud_row + 2
    _write_section_header(ws, row, "SCENARIO DESCRIPTIONS", max_col)
    for i, key in enumerate(scenario_order):
        r = row + 1 + i
        sc = model_result["scenarios"][key]["scenario"]
        ws.cell(row=r, column=1, value=sc.name).font = BOLD_VALUE
        ws.cell(row=r, column=2, value=sc.description).font = LABEL_FONT
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=max_col)

    # Interest Rate Path Explanation
    row = ud_row + 2 + len(scenario_order) + 2
    _write_section_header(ws, row, "INTEREST RATE PATH ANALYSIS", max_col)

    rate_notes = [
        ("Rising Rates Impact", "Higher WACC reduces present value of future cash flows. "
         "Companies with high debt loads are more affected. Terminal value shrinks significantly."),
        ("Falling Rates Impact", "Lower WACC increases present value of future cash flows. "
         "Growth stocks benefit disproportionately as distant cash flows gain more value."),
        ("Yield Curve Signal", "An inverted yield curve (2Y > 10Y) historically precedes recessions. "
         "Consider the bear case probability if the curve is inverted."),
    ]
    for i, (title, note) in enumerate(rate_notes):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=title).font = BOLD_VALUE
        ws.cell(row=r, column=2, value=note).font = SMALL_FONT
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=max_col)

    # ===== CHARTS =====
    chart_start = row + len(rate_notes) + 2
    cd = chart_start + 60  # data area well below charts
    ns = len(scenario_order)

    # -- Chart 1: Implied Price vs Current Price --
    ws.cell(row=cd, column=1, value="Scenario")
    ws.cell(row=cd + 1, column=1, value="Implied Price ($)")
    ws.cell(row=cd + 2, column=1, value="Current Price ($)")
    for i, key in enumerate(scenario_order):
        dcf = model_result["scenarios"][key]["dcf"]
        ws.cell(row=cd, column=2 + i, value=dcf["scenario"])
        ws.cell(row=cd + 1, column=2 + i, value=dcf["implied_share_price"])
        ws.cell(row=cd + 2, column=2 + i, value=dcf["current_price"])

    _make_bar_chart(ws, "Implied Share Price vs Current Price by Scenario", "Price ($)",
                    [cd + 1, cd + 2], cd, 2, 1 + ns,
                    ["Implied Price", "Current Price"],
                    f"A{chart_start}", width=28, height=14)

    # -- Chart 2: Enterprise Value Comparison --
    ws.cell(row=cd + 4, column=1, value="Scenario")
    ws.cell(row=cd + 5, column=1, value="PV of FCFs ($B)")
    ws.cell(row=cd + 6, column=1, value="PV of Terminal Value ($B)")
    for i, key in enumerate(scenario_order):
        dcf = model_result["scenarios"][key]["dcf"]
        ws.cell(row=cd + 4, column=2 + i, value=dcf["scenario"])
        ws.cell(row=cd + 5, column=2 + i, value=dcf["pv_fcf_total"] / 1e9)
        ws.cell(row=cd + 6, column=2 + i, value=dcf["pv_terminal_value"] / 1e9)

    chart_ev = BarChart()
    chart_ev.type = "col"
    chart_ev.grouping = "stacked"
    chart_ev.style = 10
    chart_ev.title = "Enterprise Value Breakdown by Scenario ($B)"
    chart_ev.y_axis.title = "$ Billions"
    _style_chart(chart_ev, 28, height=14)
    cats_ev = Reference(ws, min_col=2, max_col=1 + ns, min_row=cd + 4)
    d_pv = Reference(ws, min_col=2, max_col=1 + ns, min_row=cd + 5)
    d_tv = Reference(ws, min_col=2, max_col=1 + ns, min_row=cd + 6)
    chart_ev.add_data(d_pv, from_rows=True, titles_from_data=False)
    chart_ev.add_data(d_tv, from_rows=True, titles_from_data=False)
    chart_ev.set_categories(cats_ev)
    chart_ev.series[0].tx = SeriesLabel(v="PV of FCFs")
    chart_ev.series[1].tx = SeriesLabel(v="PV of Terminal Value")
    chart_ev.series[0].graphicalProperties.solidFill = MED_BLUE
    chart_ev.series[1].graphicalProperties.solidFill = ACCENT_TEAL
    ws.add_chart(chart_ev, f"A{chart_start + 15}")

    # -- Chart 3: WACC by Scenario --
    ws.cell(row=cd + 8, column=1, value="Scenario")
    ws.cell(row=cd + 9, column=1, value="WACC (%)")
    ws.cell(row=cd + 10, column=1, value="Terminal Growth (%)")
    for i, key in enumerate(scenario_order):
        dcf = model_result["scenarios"][key]["dcf"]
        ws.cell(row=cd + 8, column=2 + i, value=dcf["scenario"])
        ws.cell(row=cd + 9, column=2 + i, value=dcf["wacc"])
        ws.cell(row=cd + 10, column=2 + i, value=dcf["terminal_growth"])

    chart_wacc = BarChart()
    chart_wacc.type = "col"
    chart_wacc.style = 10
    chart_wacc.title = "WACC vs Terminal Growth Rate by Scenario"
    chart_wacc.y_axis.title = "Rate"
    chart_wacc.y_axis.numFmt = '0.0%'
    _style_chart(chart_wacc, 28, height=14)
    cats_w2 = Reference(ws, min_col=2, max_col=1 + ns, min_row=cd + 8)
    dw1 = Reference(ws, min_col=2, max_col=1 + ns, min_row=cd + 9)
    dw2 = Reference(ws, min_col=2, max_col=1 + ns, min_row=cd + 10)
    chart_wacc.add_data(dw1, from_rows=True, titles_from_data=False)
    chart_wacc.add_data(dw2, from_rows=True, titles_from_data=False)
    chart_wacc.set_categories(cats_w2)
    chart_wacc.series[0].tx = SeriesLabel(v="WACC")
    chart_wacc.series[1].tx = SeriesLabel(v="Terminal Growth")
    chart_wacc.series[0].graphicalProperties.solidFill = ACCENT_RED
    chart_wacc.series[1].graphicalProperties.solidFill = ACCENT_GREEN
    ws.add_chart(chart_wacc, f"A{chart_start + 30}")

    # -- Chart 4: FCF Projection Across All Scenarios (Line) --
    proj_years = [str(int(financials["years"][0]) + i + 1) for i in range(model_result["projection_years"])]
    n_proj = len(proj_years)
    ws.cell(row=cd + 12, column=1, value="Year")
    for j, yr in enumerate(proj_years):
        ws.cell(row=cd + 12, column=2 + j, value=yr)

    data_rows_fcf = []
    labels_fcf = []
    for si, key in enumerate(scenario_order):
        r = cd + 13 + si
        scn = model_result["scenarios"][key]
        ws.cell(row=r, column=1, value=scn["dcf"]["scenario"])
        for j, fcf_val in enumerate(scn["projection"]["projected_fcf"]):
            ws.cell(row=r, column=2 + j, value=fcf_val / 1e9)
        data_rows_fcf.append(r)
        labels_fcf.append(scn["dcf"]["scenario"])

    chart_fcf_all = _make_line_chart(
        ws, "Projected FCF Across All Scenarios ($B)", "$ Billions",
        data_rows_fcf, cd + 12, 2, 1 + n_proj, labels_fcf,
        f"A{chart_start + 45}", width=28, height=14)

    # -- Chart 5: Revenue Projection Across All Scenarios (Line) --
    ws.cell(row=cd + 19, column=1, value="Year")
    for j, yr in enumerate(proj_years):
        ws.cell(row=cd + 19, column=2 + j, value=yr)

    data_rows_rev = []
    labels_rev = []
    for si, key in enumerate(scenario_order):
        r = cd + 20 + si
        scn = model_result["scenarios"][key]
        ws.cell(row=r, column=1, value=scn["dcf"]["scenario"])
        for j, rev_val in enumerate(scn["projection"]["projected_revenue"]):
            ws.cell(row=r, column=2 + j, value=rev_val / 1e9)
        data_rows_rev.append(r)
        labels_rev.append(scn["dcf"]["scenario"])

    _make_line_chart(
        ws, "Projected Revenue Across All Scenarios ($B)", "$ Billions",
        data_rows_rev, cd + 19, 2, 1 + n_proj, labels_rev,
        f"E{chart_start + 45}", width=28, height=14)


def build_sensitivity(wb, model_result, stock, financials):
    """Sheet 12: WACC vs Terminal Growth sensitivity table."""
    ws = wb.create_sheet(title="Sensitivity Analysis")
    ws.sheet_properties.tabColor = "00695C"  # teal
    max_col = 12

    _set_col_widths(ws, {1: 24})
    for i in range(2, max_col + 1):
        ws.column_dimensions[get_column_letter(i)].width = 14

    row = 1
    _write_title_row(ws, row, "  SENSITIVITY ANALYSIS  —  IMPLIED SHARE PRICE", max_col)

    row = 3
    _write_section_header(ws, row, "WACC vs TERMINAL GROWTH RATE", max_col)

    base_wacc = model_result["wacc_data"]["wacc"]
    base_tg = model_result["terminal_growth"]
    base_dcf = model_result["scenarios"]["base"]["dcf"]

    # Terminal growth rates (columns)
    tg_range = [base_tg + delta for delta in [-0.015, -0.01, -0.005, 0, 0.005, 0.01, 0.015]]
    tg_range = [max(tg, 0.005) for tg in tg_range]

    # WACC range (rows)
    wacc_range = [base_wacc + delta for delta in [-0.03, -0.02, -0.01, 0, 0.01, 0.02, 0.03]]
    wacc_range = [max(w, 0.04) for w in wacc_range]

    # Header row
    row = 4
    ws.cell(row=row, column=1, value="WACC \\ Terminal Growth").font = HEADER_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    ws.cell(row=row, column=1).alignment = CENTER
    ws.cell(row=row, column=1).border = THIN_BORDER
    for j, tg in enumerate(tg_range):
        cell = ws.cell(row=row, column=2 + j, value=tg)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        cell.number_format = FMT_PCT2

    # Use the last projected FCF from base case for sensitivity
    last_fcf = model_result["scenarios"]["base"]["projection"]["projected_fcf"][-1]
    n_proj = model_result["projection_years"]
    net_debt = base_dcf["net_debt"]
    shares = stock["shares_outstanding"]

    for i, wacc in enumerate(wacc_range):
        r = row + 1 + i
        alt = i % 2 == 1

        # WACC label
        cell = ws.cell(row=r, column=1, value=wacc)
        cell.font = BOLD_VALUE
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        cell.number_format = FMT_PCT2
        if alt:
            cell.fill = ALT_ROW_FILL

        for j, tg in enumerate(tg_range):
            if wacc <= tg:
                val = "N/A"
                cell = ws.cell(row=r, column=2 + j, value=val)
            else:
                proj_fcfs = model_result["scenarios"]["base"]["projection"]["projected_fcf"]
                new_pv_fcfs = sum(fcf / (1 + wacc) ** (k + 1) for k, fcf in enumerate(proj_fcfs))

                term_fcf = last_fcf * (1 + tg)
                term_val = term_fcf / (wacc - tg)
                pv_term = term_val / (1 + wacc) ** n_proj
                ev = new_pv_fcfs + pv_term
                eq_val = ev - net_debt
                price = eq_val / shares if shares > 0 else 0

                cell = ws.cell(row=r, column=2 + j, value=price)
                cell.number_format = FMT_PRICE

                if price > stock["current_price"] * 1.1:
                    cell.fill = GREEN_FILL
                    cell.font = GREEN_FONT
                elif price < stock["current_price"] * 0.9:
                    cell.fill = RED_FILL
                    cell.font = RED_FONT
                else:
                    cell.font = VALUE_FONT

            cell.alignment = CENTER
            cell.border = THIN_BORDER

        # Highlight the base case row
        if abs(wacc - base_wacc) < 0.001:
            for j in range(len(tg_range)):
                ws.cell(row=r, column=2 + j).border = Border(
                    left=Side(style="thin", color=MED_GRAY),
                    right=Side(style="thin", color=MED_GRAY),
                    top=Side(style="medium", color=DARK_BLUE),
                    bottom=Side(style="medium", color=DARK_BLUE),
                )

    # Legend
    lr = row + len(wacc_range) + 2
    _write_section_header(ws, lr, "LEGEND", max_col)
    ws.cell(row=lr + 1, column=1, value="Green = >10% above current price").font = GREEN_FONT
    ws.cell(row=lr + 2, column=1, value="Red = >10% below current price").font = RED_FONT
    ws.cell(row=lr + 3, column=1, value=f"Current Price: ${stock['current_price']:.2f}").font = BOLD_VALUE
    ws.cell(row=lr + 4, column=1, value=f"Base WACC: {base_wacc:.2%}  |  Base Terminal Growth: {base_tg:.2%}").font = LABEL_FONT

    # --- Second sensitivity: Revenue Growth vs Operating Margin ---
    lr2 = lr + 6
    _write_section_header(ws, lr2, "REVENUE GROWTH vs OPERATING MARGIN SENSITIVITY", max_col)

    base_metrics = model_result["metrics"]
    base_growth = base_metrics["avg_revenue_growth"]
    base_margin = base_metrics["avg_operating_margin"]
    base_capex_pct = base_metrics["avg_capex_pct"]
    base_da_pct = base_metrics["avg_da_pct"]
    tax_rate = model_result["scenarios"]["base"]["projection"]["tax_rate"]

    growth_range = [base_growth + d for d in [-0.04, -0.02, 0, 0.02, 0.04, 0.06]]
    margin_range = [base_margin + d for d in [-0.04, -0.02, 0, 0.02, 0.04]]

    hr = lr2 + 1
    ws.cell(row=hr, column=1, value="Growth \\ Margin").font = HEADER_FONT
    ws.cell(row=hr, column=1).fill = HEADER_FILL
    ws.cell(row=hr, column=1).alignment = CENTER
    ws.cell(row=hr, column=1).border = THIN_BORDER
    for j, mg in enumerate(margin_range):
        cell = ws.cell(row=hr, column=2 + j, value=mg)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        cell.number_format = FMT_PCT

    base_rev = financials["revenue"][0]

    for i, gr in enumerate(growth_range):
        r = hr + 1 + i
        cell = ws.cell(row=r, column=1, value=gr)
        cell.font = BOLD_VALUE
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        cell.number_format = FMT_PCT

        for j, mg in enumerate(margin_range):
            proj_fcfs = []
            for yr in range(1, n_proj + 1):
                fade = 1 - (yr - 1) / (n_proj * 2)
                g = gr * fade
                rev = base_rev * (1 + g) ** yr
                ebit = rev * max(mg, 0.01)
                nopat = ebit * (1 - tax_rate)
                da = rev * base_da_pct
                capex = rev * base_capex_pct
                fcf = nopat + da - capex
                proj_fcfs.append(fcf)

            w = base_wacc
            new_pv = sum(fcf / (1 + w) ** (k + 1) for k, fcf in enumerate(proj_fcfs))
            tf = proj_fcfs[-1] * (1 + base_tg)
            tv = tf / (w - base_tg) if w > base_tg else 0
            pvt = tv / (1 + w) ** n_proj
            ev = new_pv + pvt
            eq = ev - net_debt
            price = eq / shares if shares > 0 else 0

            cell = ws.cell(row=r, column=2 + j, value=price)
            cell.number_format = FMT_PRICE
            cell.alignment = CENTER
            cell.border = THIN_BORDER

            if price > stock["current_price"] * 1.1:
                cell.fill = GREEN_FILL
                cell.font = GREEN_FONT
            elif price < stock["current_price"] * 0.9:
                cell.fill = RED_FILL
                cell.font = RED_FONT

    # ===== CHARTS for Sensitivity =====
    sens_chart_row = hr + len(growth_range) + 3
    scd = sens_chart_row + 32  # data area

    # -- Chart: Impact of WACC on Implied Price (line) --
    ws.cell(row=scd, column=1, value="WACC")
    ws.cell(row=scd + 1, column=1, value="Implied Price ($)")
    mid_tg_idx = len(tg_range) // 2  # use middle terminal growth
    for i, wacc in enumerate(wacc_range):
        ws.cell(row=scd, column=2 + i, value=f"{wacc:.1%}")
        # get the price from the sensitivity table
        r = 5 + i  # row where the data was written
        c = 2 + mid_tg_idx
        cell_val = ws.cell(row=r, column=c).value
        ws.cell(row=scd + 1, column=2 + i, value=cell_val if isinstance(cell_val, (int, float)) else 0)

    chart_wacc_sens = _make_line_chart(
        ws, "Impact of WACC on Implied Share Price", "Implied Price ($)",
        [scd + 1], scd, 2, 1 + len(wacc_range),
        ["Implied Price at Base Terminal Growth"],
        f"A{sens_chart_row}", width=28, height=14)

    # -- Chart: Impact of Terminal Growth on Implied Price (line) --
    ws.cell(row=scd + 3, column=1, value="Terminal Growth")
    ws.cell(row=scd + 4, column=1, value="Implied Price ($)")
    mid_wacc_idx = len(wacc_range) // 2  # use middle WACC (base)
    for j, tg in enumerate(tg_range):
        ws.cell(row=scd + 3, column=2 + j, value=f"{tg:.2%}")
        r = 5 + mid_wacc_idx
        c = 2 + j
        cell_val = ws.cell(row=r, column=c).value
        ws.cell(row=scd + 4, column=2 + j, value=cell_val if isinstance(cell_val, (int, float)) else 0)

    _make_line_chart(
        ws, "Impact of Terminal Growth on Implied Share Price", "Implied Price ($)",
        [scd + 4], scd + 3, 2, 1 + len(tg_range),
        ["Implied Price at Base WACC"],
        f"A{sens_chart_row + 15}", width=28, height=14)


# ============================================================================
# MAIN BUILDER
# ============================================================================

def build_workbook(stock, financials, rates, model_result, data_source) -> Workbook:
    """Build the complete DCF workbook and return the Workbook object."""
    wb = Workbook()

    # Sheet 1: Dashboard
    build_dashboard(wb, stock, financials, rates, model_result, data_source)

    # Sheet 2-4: Financial Statements
    build_income_statement(wb, financials)
    build_balance_sheet(wb, financials)
    build_cash_flow(wb, financials)

    # Sheet 5: WACC
    build_wacc_sheet(wb, model_result["wacc_data"], stock, financials, rates)

    # Sheet 6-10: DCF Scenarios
    for key in ["base", "bull", "bear", "rate_hike", "rate_cut"]:
        build_dcf_scenario_sheet(wb, key, model_result, financials, stock)

    # Sheet 11: Scenario Comparison
    build_scenario_comparison(wb, model_result, stock, financials)

    # Sheet 12: Sensitivity Analysis
    build_sensitivity(wb, model_result, stock, financials)

    return wb


def save_workbook(wb: Workbook, filepath: str):
    """Save workbook and print confirmation."""
    wb.save(filepath)
    print(f"[excel_builder] Workbook saved to: {filepath}")
